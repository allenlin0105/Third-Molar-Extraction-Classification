import math
import shutil
import argparse
from pathlib import Path

import matplotlib.pyplot as plt
from PIL import Image
from tqdm import trange
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from COCOParser import COCOParser

def copy_images(source_folder, target_folder, filenames=[]):
    target_folder.mkdir(parents=True, exist_ok=True)

    if len(filenames) == 0:
        filenames = [image_path.name for image_path in source_folder.iterdir()]
    for filename in filenames:
        shutil.copyfile(source_folder.joinpath(filename), target_folder.joinpath(filename))

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dataset_path", type=Path, default="data/odontoai-v3")
    args = parser.parse_args()

    dataset_path = args.dataset_path
    new_dataset_path = dataset_path.parents[0].joinpath("odontoai-labeling")

    target_tooth_names = ["tooth-18", "tooth-28", "tooth-38", "tooth-48"]

    splits = ["train", "val"]
    
    """
    A list of lists: (file_path, bboxes)
    bboxes: a dict(category_name: bbox)
    """
    file_path_bboxes_list = []
    for split in splits:
        split_folder = dataset_path.joinpath(split)
        coco_annotations_file = split_folder.joinpath(f"{split}.json")
        coco = COCOParser(coco_annotations_file)

        image_ids = coco.get_image_ids()
        for image_id in image_ids:
            image_filename = coco.get_image_filename(image_id)
            file_path_bboxes_list.append([
                split_folder.joinpath("images", image_filename),
                {}
            ])

            anns = coco.get_anns(image_id)
            for ann in anns:
                category_name = coco.get_category_name(ann["category_id"])
                if category_name not in target_tooth_names:
                    continue
                file_path_bboxes_list[-1][1][category_name] = ann

    print(f"Total number of images to be labeled: {len(file_path_bboxes_list)}")

    # Sort by file names
    file_path_bboxes_list.sort(key=lambda x: x[0])

    n_group = 6
    for group_index in range(n_group):
        group_folder = new_dataset_path.joinpath(f"group_{group_index}")
        group_folder.mkdir(exist_ok=True, parents=True)

        wb = Workbook()
        ws = wb.active
        
        # Set header row
        ws["A1"] = "file"
        ws["B1"] = "tooth"
        ws["C1"] = "direction"
        ws["D1"] = "method"

        # Set dropdown list
        direction_dv = DataValidation(type="list", formula1='"直,橫"', allow_blank=True)
        method_dv = DataValidation(type="list", formula1='"複雜拔牙,單純齒切,複雜齒切"', allow_blank=True)
        
        ws.add_data_validation(direction_dv)
        ws.add_data_validation(method_dv)

        excel_row_index = 2
        for file_index in trange(group_index, len(file_path_bboxes_list), n_group, desc=f"Group {group_index}"):
            # Common used
            source_image_file_path = file_path_bboxes_list[file_index][0]
            bboxes = file_path_bboxes_list[file_index][1]

            # Excel
            for target_tooth_name in target_tooth_names:
                if target_tooth_name not in bboxes.keys():
                    continue

                ws[f"A{excel_row_index}"] = source_image_file_path.name
                ws[f"B{excel_row_index}"] = int(target_tooth_name.split("-")[1])

                excel_row_index += 1

            # Image
            target_image_file_path = group_folder.joinpath(source_image_file_path.name)

            fig, ax = plt.subplots(figsize=(15,10))
            for category_name, bbox in bboxes.items():
                x, y, w, h = [int(b) for b in bbox["bbox"]]
                rect = plt.Rectangle((x, y), w, h, linewidth=2, edgecolor="red", facecolor='none')
                ax.add_patch(rect)

                t_box = ax.text(x, y, category_name, color='black', fontsize=8)
                t_box.set_bbox(dict(boxstyle='square, pad=0.05', facecolor='white', alpha=0.6))

            image = Image.open(source_image_file_path)
            ax.axis('off')
            ax.imshow(image)
            ax.set_xlabel('Longitude')

            plt.tight_layout()
            plt.savefig(target_image_file_path.__str__())
            plt.clf()
            plt.close()

        # Add cell to dropdown list
        direction_dv.add(f"C2:C{excel_row_index - 1}")
        method_dv.add(f"D2:D{excel_row_index - 1}")

        # Save excel file
        wb.save(group_folder.joinpath('label.xlsx'))

if __name__=="__main__":
    main()